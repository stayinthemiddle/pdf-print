#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import shutil
import time
import logging
import re
from datetime import datetime
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import PyPDF2
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pdf_manager.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class PDFMetadataExtractor:
    """PDF元数据提取器"""
    
    @staticmethod
    def extract_metadata(pdf_path):
        """从PDF中提取元数据"""
        metadata = {
            'title': '',
            'author': '',
            'journal': '',
            'year': '',
            'doi': ''
        }
        
        try:
            # 使用PyPDF2提取基本元数据
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                info = reader.metadata
                
                if info:
                    metadata['title'] = info.get('/Title', '') or ''
                    metadata['author'] = info.get('/Author', '') or ''
                    
                    # 尝试从Subject或Keywords中提取期刊信息
                    subject = info.get('/Subject', '') or ''
                    keywords = info.get('/Keywords', '') or ''
                    
                    # 简单的年份提取
                    creation_date = info.get('/CreationDate', '')
                    if creation_date and len(creation_date) >= 4:
                        year_match = re.search(r'(\d{4})', str(creation_date))
                        if year_match:
                            metadata['year'] = year_match.group(1)
            
            # 使用pdfplumber提取文本中的DOI
            with pdfplumber.open(pdf_path) as pdf:
                # 只检查前两页以提高性能
                for i, page in enumerate(pdf.pages[:2]):
                    text = page.extract_text() or ''
                    
                    # 提取DOI
                    doi_pattern = r'(?:DOI|doi)[\s:]*([^\s]+)'
                    doi_match = re.search(doi_pattern, text)
                    if doi_match:
                        metadata['doi'] = doi_match.group(1).strip()
                    
                    # 如果标题为空，尝试从第一页提取
                    if not metadata['title'] and i == 0:
                        lines = text.split('\n')
                        # 假设标题在前几行
                        for line in lines[:5]:
                            line = line.strip()
                            if len(line) > 10 and not any(char.isdigit() for char in line[:5]):
                                metadata['title'] = line
                                break
                    
                    # 尝试提取期刊信息
                    if not metadata['journal']:
                        journal_patterns = [
                            r'(?:Journal of|IEEE|Nature|Science|Cell|Proceedings of)[^,\n]+',
                            r'(?:International|American|European|Asian) Journal of[^,\n]+'
                        ]
                        for pattern in journal_patterns:
                            journal_match = re.search(pattern, text, re.IGNORECASE)
                            if journal_match:
                                metadata['journal'] = journal_match.group(0).strip()
                                break
                
        except Exception as e:
            logging.error(f"提取PDF元数据时出错 {pdf_path}: {str(e)}")
        
        return metadata

class ExcelManager:
    """Excel文件管理器"""
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.columns = ['序号', '文件名', '原始文件名', '类型', '标题', '作者', '期刊', '年份', 'DOI', '添加时间']
        self._ensure_excel_exists()
    
    def _ensure_excel_exists(self):
        """确保Excel文件存在，如果不存在则创建"""
        if not os.path.exists(self.excel_path):
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.excel_path, index=False, engine='openpyxl')
            logging.info(f"创建新的Excel文件: {self.excel_path}")
    
    def add_record(self, file_info):
        """添加新记录到Excel"""
        try:
            # 读取现有数据
            df = pd.read_excel(self.excel_path, engine='openpyxl')
            
            # 创建新记录
            new_record = pd.DataFrame([file_info])
            
            # 添加到DataFrame
            df = pd.concat([df, new_record], ignore_index=True)
            
            # 保存到Excel
            df.to_excel(self.excel_path, index=False, engine='openpyxl')
            
            # 自动调整列宽
            self._auto_adjust_columns()
            
            logging.info(f"成功添加记录: {file_info['文件名']}")
            
        except Exception as e:
            logging.error(f"添加Excel记录时出错: {str(e)}")
    
    def _auto_adjust_columns(self):
        """自动调整Excel列宽"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(self.excel_path)
            
        except Exception as e:
            logging.warning(f"调整列宽时出错: {str(e)}")

class PDFHandler(FileSystemEventHandler):
    """PDF文件处理器"""
    
    def __init__(self, watch_dirs, excel_manager):
        self.watch_dirs = watch_dirs
        self.excel_manager = excel_manager
        self.chinese_counter = self._get_max_counter('中文pdf', 'c')
        self.english_counter = self._get_max_counter('英文pdf', 'e')
    
    def _get_max_counter(self, folder, prefix):
        """获取文件夹中最大的计数器值"""
        folder_path = self.watch_dirs.get(folder)
        if not folder_path or not os.path.exists(folder_path):
            return 0
        
        max_num = 0
        pattern = re.compile(f'^{prefix}(\d+)\.pdf$')
        
        for filename in os.listdir(folder_path):
            match = pattern.match(filename)
            if match:
                num = int(match.group(1))
                max_num = max(max_num, num)
        
        return max_num
    
    def on_created(self, event):
        """文件创建时触发"""
        if event.is_directory:
            return
        
        if event.src_path.lower().endswith('.pdf'):
            # 检查文件是否已经是目标格式
            filename = os.path.basename(event.src_path)
            if re.match(r'^[ce]\d+\.pdf$', filename):
                return
            # 等待文件写入完成
            time.sleep(0.5)
            self.process_pdf(event.src_path)
    
    def on_moved(self, event):
        """文件移动时触发（拖入）"""
        if event.is_directory:
            return
        
        if event.dest_path.lower().endswith('.pdf'):
            # 检查文件是否已经是目标格式
            filename = os.path.basename(event.dest_path)
            if re.match(r'^[ce]\d+\.pdf$', filename):
                return
            # 等待文件移动完成
            time.sleep(0.5)
            self.process_pdf(event.dest_path)
    
    def process_pdf(self, file_path):
        """处理PDF文件"""
        try:
            # 确定文件类型
            parent_dir = os.path.basename(os.path.dirname(file_path))
            
            if parent_dir == '中文pdf':
                file_type = '中文'
                prefix = 'c'
                self.chinese_counter += 1
                counter = self.chinese_counter
            elif parent_dir == '英文pdf':
                file_type = '英文'
                prefix = 'e'
                self.english_counter += 1
                counter = self.english_counter
            else:
                logging.warning(f"未知的文件夹: {parent_dir}")
                return
            
            # 生成新文件名
            new_filename = f"{prefix}{counter:02d}.pdf"
            new_path = os.path.join(os.path.dirname(file_path), new_filename)
            
            # 保存原始文件名
            original_filename = os.path.basename(file_path)
            
            # 如果文件已经是目标格式，跳过
            if file_path == new_path:
                logging.info(f"文件已经是目标格式: {new_filename}")
                return
            
            # 重命名文件
            os.rename(file_path, new_path)
            logging.info(f"重命名: {original_filename} -> {new_filename}")
            
            # 提取元数据
            metadata = PDFMetadataExtractor.extract_metadata(new_path)
            
            # 准备Excel记录
            record = {
                '序号': counter,
                '文件名': new_filename,
                '原始文件名': original_filename,
                '类型': file_type,
                '标题': metadata['title'],
                '作者': metadata['author'],
                '期刊': metadata['journal'],
                '年份': metadata['year'],
                'DOI': metadata['doi'],
                '添加时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # 添加到Excel
            self.excel_manager.add_record(record)
            
        except Exception as e:
            logging.error(f"处理PDF文件时出错 {file_path}: {str(e)}")

class PDFManager:
    """PDF管理器主类"""
    
    def __init__(self):
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.chinese_dir = os.path.join(self.base_dir, '中文pdf')
        self.english_dir = os.path.join(self.base_dir, '英文pdf')
        self.excel_path = os.path.join(self.base_dir, 'pdf_records.xlsx')
        
        # 确保文件夹存在
        self._ensure_directories()
        
        # 初始化Excel管理器
        self.excel_manager = ExcelManager(self.excel_path)
        
        # 初始化监控器
        self.observer = Observer()
        
    def _ensure_directories(self):
        """确保必要的文件夹存在"""
        for dir_path in [self.chinese_dir, self.english_dir]:
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
                logging.info(f"创建文件夹: {dir_path}")
    
    def start(self):
        """启动PDF管理器"""
        logging.info("PDF管理系统启动...")
        logging.info(f"监控文件夹: {self.chinese_dir}")
        logging.info(f"监控文件夹: {self.english_dir}")
        logging.info(f"Excel记录文件: {self.excel_path}")
        
        # 创建事件处理器
        watch_dirs = {
            '中文pdf': self.chinese_dir,
            '英文pdf': self.english_dir
        }
        handler = PDFHandler(watch_dirs, self.excel_manager)
        
        # 设置监控
        self.observer.schedule(handler, self.chinese_dir, recursive=False)
        self.observer.schedule(handler, self.english_dir, recursive=False)
        
        # 启动监控
        self.observer.start()
        
        try:
            print("\n✅ PDF管理系统已启动！")
            print("📁 监控文件夹:")
            print(f"   - 中文PDF: {self.chinese_dir}")
            print(f"   - 英文PDF: {self.english_dir}")
            print("📊 Excel记录: pdf_records.xlsx")
            print("\n🔄 正在监控文件夹，将PDF文件拖入相应文件夹即可自动处理...")
            print("📝 按 Ctrl+C 停止程序\n")
            
            while True:
                time.sleep(1)
                
        except KeyboardInterrupt:
            print("\n正在停止程序...")
            self.observer.stop()
            logging.info("PDF管理系统已停止")
        
        self.observer.join()

def main():
    """主函数"""
    manager = PDFManager()
    manager.start()

if __name__ == "__main__":
    main()