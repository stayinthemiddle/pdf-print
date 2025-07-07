#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PDF批量处理程序 - 手动运行版本
每次运行时处理所有未处理的PDF文件
支持AI增强的元数据提取
"""

import os
import re
import logging
from datetime import datetime
from pathlib import Path
import argparse
from typing import Dict, Optional

import PyPDF2
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# 导入AI模块（如果可用）
try:
    from llm_extractor import LLMExtractor
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False
    print("警告: AI模块未安装，将使用传统提取方法")

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pdf_batch.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class PDFMetadataExtractor:
    """PDF元数据提取器"""
    
    @staticmethod
    def extract_metadata(pdf_path):
        """从PDF中提取元数据"""
        metadata = {
            'title': '',
            'author': '',
            'journal': '',
            'year': '',
            'doi': ''
        }
        
        try:
            # 使用PyPDF2提取基本元数据
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                info = reader.metadata
                
                if info:
                    metadata['title'] = info.get('/Title', '') or ''
                    metadata['author'] = info.get('/Author', '') or ''
                    
                    # 简单的年份提取
                    creation_date = info.get('/CreationDate', '')
                    if creation_date and len(str(creation_date)) >= 4:
                        year_match = re.search(r'(\d{4})', str(creation_date))
                        if year_match:
                            metadata['year'] = year_match.group(1)
            
            # 使用pdfplumber提取文本中的DOI
            with pdfplumber.open(pdf_path) as pdf:
                # 只检查前两页以提高性能
                for i, page in enumerate(pdf.pages[:2]):
                    text = page.extract_text() or ''
                    
                    # 提取DOI
                    doi_pattern = r'(?:DOI|doi)[\s:]*([^\s]+)'
                    doi_match = re.search(doi_pattern, text)
                    if doi_match:
                        metadata['doi'] = doi_match.group(1).strip()
                    
                    # 如果标题为空，尝试从第一页提取
                    if not metadata['title'] and i == 0:
                        lines = text.split('\n')
                        # 假设标题在前几行
                        for line in lines[:5]:
                            line = line.strip()
                            if len(line) > 10 and not any(char.isdigit() for char in line[:5]):
                                metadata['title'] = line
                                break
                    
                    # 尝试提取期刊信息
                    if not metadata['journal']:
                        journal_patterns = [
                            r'(?:Journal of|IEEE|Nature|Science|Cell|Proceedings of)[^,\n]+',
                            r'(?:International|American|European|Asian) Journal of[^,\n]+'
                        ]
                        for pattern in journal_patterns:
                            journal_match = re.search(pattern, text, re.IGNORECASE)
                            if journal_match:
                                metadata['journal'] = journal_match.group(0).strip()
                                break
                
        except Exception as e:
            logging.error(f"提取PDF元数据时出错 {pdf_path}: {str(e)}")
        
        return metadata

class ExcelManager:
    """Excel文件管理器"""
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.columns = ['序号', '文件名', '原始文件名', '类型', '标题', '作者', '期刊', '年份', 'DOI', 
                       '添加时间', '提取方式', '提取置信度', '配对文献', '配对置信度']
        self._ensure_excel_exists()
    
    def _ensure_excel_exists(self):
        """确保Excel文件存在，如果不存在则创建"""
        if not os.path.exists(self.excel_path):
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.excel_path, index=False, engine='openpyxl')
            logging.info(f"创建新的Excel文件: {self.excel_path}")
    
    def add_record(self, file_info):
        """添加新记录到Excel"""
        try:
            # 读取现有数据
            df = pd.read_excel(self.excel_path, engine='openpyxl')
            
            # 创建新记录
            new_record = pd.DataFrame([file_info])
            
            # 添加到DataFrame
            df = pd.concat([df, new_record], ignore_index=True)
            
            # 保存到Excel
            df.to_excel(self.excel_path, index=False, engine='openpyxl')
            
            # 自动调整列宽
            self._auto_adjust_columns()
            
            logging.info(f"成功添加记录: {file_info['文件名']}")
            
        except Exception as e:
            logging.error(f"添加Excel记录时出错: {str(e)}")
    
    def _auto_adjust_columns(self):
        """自动调整Excel列宽"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(self.excel_path)
            
        except Exception as e:
            logging.warning(f"调整列宽时出错: {str(e)}")

class PDFBatchProcessor:
    """PDF批量处理器"""
    
    def __init__(self, use_ai=False):
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.chinese_dir = os.path.join(self.base_dir, '中文pdf')
        self.english_dir = os.path.join(self.base_dir, '英文pdf')
        self.excel_path = os.path.join(self.base_dir, 'pdf_records.xlsx')
        
        # 确保文件夹存在
        self._ensure_directories()
        
        # 初始化Excel管理器
        self.excel_manager = ExcelManager(self.excel_path)
        
        # 初始化计数器
        self.chinese_counter = self._get_max_counter('中文pdf', 'c')
        self.english_counter = self._get_max_counter('英文pdf', 'e')
        
        # AI提取器
        self.use_ai = use_ai and AI_AVAILABLE
        if self.use_ai:
            self.llm_extractor = LLMExtractor()
            logging.info("已启用AI元数据提取")
        else:
            self.traditional_extractor = PDFMetadataExtractor()
    
    def _ensure_directories(self):
        """确保必要的文件夹存在"""
        for dir_path in [self.chinese_dir, self.english_dir]:
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
                logging.info(f"创建文件夹: {dir_path}")
    
    def _get_max_counter(self, folder_name, prefix):
        """获取文件夹中最大的计数器值"""
        folder_path = os.path.join(self.base_dir, folder_name)
        if not os.path.exists(folder_path):
            return 0
        
        max_num = 0
        pattern = re.compile(f'^{prefix}(\d+)\.pdf$')
        
        for filename in os.listdir(folder_path):
            match = pattern.match(filename)
            if match:
                num = int(match.group(1))
                max_num = max(max_num, num)
        
        return max_num
    
    def process_folder(self, folder_path, file_type, prefix):
        """处理一个文件夹中的所有PDF文件"""
        processed_count = 0
        
        if not os.path.exists(folder_path):
            return processed_count
        
        # 获取所有PDF文件
        pdf_files = [f for f in os.listdir(folder_path) 
                     if f.lower().endswith('.pdf') and not re.match(f'^{prefix}\d+\.pdf$', f)]
        
        # 使用进度条
        for filename in tqdm(pdf_files, desc=f"处理{file_type}PDF", disable=len(pdf_files) == 0):
            file_path = os.path.join(folder_path, filename)
            
            # 更新计数器
            if file_type == '中文':
                self.chinese_counter += 1
                counter = self.chinese_counter
            else:
                self.english_counter += 1
                counter = self.english_counter
            
            # 生成新文件名
            new_filename = f"{prefix}{counter:02d}.pdf"
            new_path = os.path.join(folder_path, new_filename)
            
            try:
                # 重命名文件
                os.rename(file_path, new_path)
                logging.info(f"重命名: {filename} -> {new_filename}")
                
                # 提取元数据
                if self.use_ai:
                    metadata = self.llm_extractor.extract_metadata(new_path, use_llm=True)
                    extraction_method = 'AI'
                    confidence = metadata.get('confidence', '0')
                else:
                    metadata = self.traditional_extractor.extract_metadata(new_path)
                    extraction_method = '传统'
                    confidence = '30'  # 传统方法默认置信度
                
                # 准备Excel记录
                record = {
                    '序号': counter,
                    '文件名': new_filename,
                    '原始文件名': filename,
                    '类型': file_type,
                    '标题': metadata.get('title', ''),
                    '作者': metadata.get('authors', metadata.get('author', '')),
                    '期刊': metadata.get('journal', ''),
                    '年份': metadata.get('year', ''),
                    'DOI': metadata.get('doi', ''),
                    '添加时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '提取方式': extraction_method,
                    '提取置信度': confidence,
                    '配对文献': '',
                    '配对置信度': ''
                }
                
                # 添加到Excel
                self.excel_manager.add_record(record)
                processed_count += 1
                
            except Exception as e:
                logging.error(f"处理文件 {filename} 时出错: {str(e)}")
        
        return processed_count
    
    def run(self):
        """运行批量处理"""
        print("\n🚀 PDF批量处理程序")
        if self.use_ai:
            print("🤖 已启用AI元数据提取")
        print("=" * 50)
        
        # 处理中文PDF
        print("\n📂 处理中文PDF文件夹...")
        chinese_count = self.process_folder(self.chinese_dir, '中文', 'c')
        print(f"   ✅ 处理了 {chinese_count} 个中文PDF文件")
        
        # 处理英文PDF
        print("\n📂 处理英文PDF文件夹...")
        english_count = self.process_folder(self.english_dir, '英文', 'e')
        print(f"   ✅ 处理了 {english_count} 个英文PDF文件")
        
        # 总结
        total_count = chinese_count + english_count
        print("\n" + "=" * 50)
        print(f"📊 处理完成！总共处理了 {total_count} 个文件")
        
        if total_count > 0:
            print(f"📝 所有记录已保存到: pdf_records.xlsx")
        else:
            print("💡 没有发现需要处理的新文件")
            print("   提示：请确保PDF文件放在正确的文件夹中")
            print("   - 中文PDF -> 中文pdf/")
            print("   - 英文PDF -> 英文pdf/")

def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='PDF批量处理程序')
    parser.add_argument('--use-ai', action='store_true', help='使用AI提取元数据')
    parser.add_argument('--batch-size', type=int, default=20, help='批处理大小')
    
    args = parser.parse_args()
    
    processor = PDFBatchProcessor(use_ai=args.use_ai)
    processor.run()

if __name__ == "__main__":
    main()